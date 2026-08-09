import base64
import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from rifthelper import config

# --- Paleta de marca ---
NAVY = "0A1428"
NAVY_2 = "14264D"
GOLD = "C89B3C"
GOLD_SOFT = "F0C868"
LIGHT = "F2F5FA"
BORDER = "D5DAE4"
WHITE = "FFFFFF"
MUTED = "C7D2E8"
GREEN = "1E8449"
RED = "C0392B"

FONT_NAME = "Segoe UI"

THIN = Side(style="thin", color=BORDER)


def generate_stats_excel(
    full_name: str, rank_icon_path: Path | None, matches: list[dict], mode_label: str = ""
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas"
    ws.sheet_view.showGridLines = False

    total = len(matches)
    wins = sum(1 for m in matches if m.get("win"))
    wr = round(wins / total * 100) if total else 0

    include_boots = any(m.get("is_adc") for m in matches)
    total_cols = 22 if include_boots else 21

    # ================= PANEL DEL JUGADOR =================
    for row in range(1, 5):
        for col in range(1, total_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=NAVY_2)
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor=NAVY_2)

    if rank_icon_path and rank_icon_path.is_file():
        img = XLImage(str(rank_icon_path))
        img.width = 52
        img.height = 52
        ws.add_image(img, "A2")

    ws["C2"] = full_name
    ws["C2"].font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    ws["C3"] = f"Partidas analizadas: {total}" + (f" · {mode_label}" if mode_label else "")
    ws["C3"].font = Font(name=FONT_NAME, size=11, color=MUTED)
    ws["C4"] = f"Winrate total: {wr}%  ({wins}V / {total - wins}D)"
    ws["C4"].font = Font(name=FONT_NAME, size=12, bold=True, color=GOLD_SOFT)

    # ================= CABECERA =================
    header_row = 6
    headers = [
        "Nº", "Fecha", "Campeón", "Adversario", "Oro diff @10", "Oro diff @30", "CS/min",
        "Daño/min", "Daño total", "KP%", "Daño edif.", "Visión/min", "Runa",
        "Item 1", "Item 2", "Item 3", "Item 4", "Item 5", "Item 6",
    ]
    if include_boots:
        headers.append("Botas")
    headers += ["Trinket", "Resultado"]

    ws.row_dimensions[header_row].height = 26
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=GOLD))

    widths = [6, 12, 6, 6, 13, 13, 9, 10, 12, 8, 12, 11, 6, 6, 6, 6, 6, 6, 6]
    if include_boots:
        widths.append(6)
    widths += [6, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ================= FILAS =================
    row = header_row + 1
    first_data = row
    for i, m in enumerate(matches, 1):
        ws.row_dimensions[row].height = 30
        zebra = PatternFill("solid", fgColor=LIGHT) if i % 2 == 0 else None

        values = [
            i,
            m.get("date", "—"),
            "",
            "",
            _fmt_diff(m.get("diff10")),
            _fmt_diff(m.get("diff30")),
            m.get("cs_min", 0),
            m.get("dmg_per_min", 0),
            m.get("total_damage", 0),
            f"{m.get('kp', 0)}%",
            m.get("dmg_buildings", 0),
            m.get("vision_per_min", 0),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        if include_boots:
            values.append("")
        values += ["", "Victoria" if m.get("win") else "Derrota"]

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=THIN, right=THIN, bottom=THIN)
            if zebra:
                cell.fill = zebra

        trinket_col = 21 if include_boots else 20
        result_col = trinket_col + 1
        res = ws.cell(row=row, column=result_col)
        res.font = Font(name=FONT_NAME, size=10, bold=True, color=GREEN if m.get("win") else RED)

        _add_cell_image(ws, m.get("player_icon"), f"C{row}")
        _add_cell_image(ws, m.get("enemy_icon"), f"D{row}")
        _add_cell_image(ws, _rune_icon(m.get("keystone")), f"M{row}")
        items = m.get("item_ids", []) or []
        for slot in range(6):
            if slot < len(items):
                _add_cell_image(ws, _item_icon(items[slot]), f"{get_column_letter(14 + slot)}{row}")
        if include_boots and m.get("is_adc") and len(items) > 7:
            _add_cell_image(ws, _item_icon(items[7]), f"T{row}")
        if len(items) > 6:
            _add_cell_image(ws, _item_icon(items[6]), f"{get_column_letter(trinket_col)}{row}")
        row += 1

    # ================= ACABADO =================
    ws.freeze_panes = "A7"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fmt_diff(value):
    if value is None:
        return "—"
    return f"+{value}" if value > 0 else str(value)


def _add_cell_image(ws, path, anchor: str, size: int = 40) -> None:
    if not path or not Path(path).is_file():
        return
    img = XLImage(str(path))
    img.width = size
    img.height = size
    ws.add_image(img, anchor)


def _item_icon(item_id) -> Path | None:
    if not item_id:
        return None
    path = config.ITEMS_DIR / f"{item_id}.png"
    return path if path.is_file() else None


def _rune_icon(rune_id) -> Path | None:
    if not rune_id:
        return None
    path = config.RUNES_DIR / f"{rune_id}.png"
    return path if path.is_file() else None


def _data_uri(path: Path | None, size: int = 40) -> str:
    """Devuelve un <img> con la imagen embebida en base64, o una celda vacía."""
    if not path or not path.is_file():
        return '<div class="slot"></div>'
    b64 = base64.b64encode(path.read_bytes()).decode("ascii")
    return f'<img class="cell-img" style="width:{size}px;height:{size}px" src="data:image/png;base64,{b64}" alt="">'


def generate_stats_html(
    full_name: str, rank_icon_path: Path | None, matches: list[dict], mode_label: str = ""
) -> str:
    total = len(matches)
    wins = sum(1 for m in matches if m.get("win"))
    wr = round(wins / total * 100) if total else 0
    include_boots = any(m.get("is_adc") for m in matches)

    headers = [
        "Nº", "Fecha", "Campeón", "Adversario", "Oro diff @10", "Oro diff @30", "CS/min",
        "Daño/min", "Daño total", "KP%", "Daño edif.", "Visión/min", "Runa",
        "Item 1", "Item 2", "Item 3", "Item 4", "Item 5", "Item 6",
    ]
    if include_boots:
        headers.append("Botas")
    headers += ["Trinket", "Resultado"]

    thead = "".join(f"<th>{h}</th>" for h in headers)

    body = []
    for i, m in enumerate(matches, 1):
        items = m.get("item_ids", []) or []
        result = "Victoria" if m.get("win") else "Derrota"
        res_cls = "win" if m.get("win") else "loss"

        cells = [
            f"<td>{i}</td>",
            f"<td>{m.get('date', '—')}</td>",
            f"<td>{_data_uri(m.get('player_icon'))}</td>",
            f"<td>{_data_uri(m.get('enemy_icon'))}</td>",
            f"<td>{_fmt_diff(m.get('diff10'))}</td>",
            f"<td>{_fmt_diff(m.get('diff30'))}</td>",
            f"<td>{m.get('cs_min', 0)}</td>",
            f"<td>{m.get('dmg_per_min', 0)}</td>",
            f"<td>{m.get('total_damage', 0)}</td>",
            f"<td>{m.get('kp', 0)}%</td>",
            f"<td>{m.get('dmg_buildings', 0)}</td>",
            f"<td>{m.get('vision_per_min', 0)}</td>",
            f"<td>{_data_uri(_rune_icon(m.get('keystone')))}</td>",
        ]
        for slot in range(6):
            cells.append(f"<td>{_data_uri(_item_icon(items[slot]) if slot < len(items) else None)}</td>")
        if include_boots:
            boots = items[7] if (m.get("is_adc") and len(items) > 7) else None
            cells.append(f"<td>{_data_uri(_item_icon(boots))}</td>")
        trinket = items[6] if len(items) > 6 else None
        cells.append(f"<td>{_data_uri(_item_icon(trinket))}</td>")
        cells.append(f'<td class="{res_cls}">{result}</td>')

        zebra = ' class="zebra"' if i % 2 == 0 else ""
        body.append(f"<tr{zebra}>" + "".join(cells) + "</tr>")

    rank_img = _data_uri(rank_icon_path, size=52)
    mode_txt = f" · {mode_label}" if mode_label else ""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Stats · {full_name}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; padding: 24px; background: #0A1428; color: #F2F5FA;
         font-family: 'Segoe UI', Arial, sans-serif; }}
  .panel {{ background: #14264D; border: 1px solid #C89B3C; border-radius: 10px;
            padding: 16px 20px; margin-bottom: 18px; display: flex; align-items: center; gap: 16px; }}
  .panel img {{ border-radius: 6px; }}
  .panel h1 {{ margin: 0; font-size: 22px; }}
  .panel .sub {{ color: #C7D2E8; font-size: 13px; margin: 2px 0; }}
  .panel .wr {{ color: #F0C868; font-weight: 700; font-size: 15px; }}
  .wrap {{ overflow-x: auto; border: 1px solid #2b3a5e; border-radius: 8px; }}
  table {{ border-collapse: collapse; width: 100%; background: #FFFFFF; color: #1a2438;
          font-size: 13px; min-width: 1280px; }}
  th {{ background: #0A1428; color: #fff; font-weight: 700; padding: 8px 6px;
       border-bottom: 3px solid #C89B3C; white-space: nowrap; }}
  td {{ padding: 4px 6px; text-align: center; border-left: 1px solid #D5DAE4;
       border-bottom: 1px solid #D5DAE4; }}
  tr.zebra td {{ background: #F2F5FA; }}
  .cell-img {{ display: block; margin: 0 auto; border-radius: 4px; }}
  .slot {{ width: 40px; height: 40px; margin: 0 auto; }}
  td.win {{ color: #1E8449; font-weight: 700; }}
  td.loss {{ color: #C0392B; font-weight: 700; }}
</style>
</head>
<body>
  <div class="panel">
    {rank_img}
    <div>
      <h1>{full_name}</h1>
      <div class="sub">Partidas analizadas: {total}{mode_txt}</div>
      <div class="wr">Winrate total: {wr}% ({wins}V / {total - wins}D)</div>
    </div>
  </div>
  <div class="wrap">
    <table>
      <thead><tr>{thead}</tr></thead>
      <tbody>{''.join(body)}</tbody>
    </table>
  </div>
</body>
</html>"""

